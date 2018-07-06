import openpyxl 

PostosVazoes = []

planilhas = openpyxl.load_workbook('example.xlsx')
nomes_das_abas = planilhas.get_sheet_names()
print(nomes_das_abas)

for abas in nomes_das_abas:
    planilha = wb.get_sheet_by_name(abas)
    NumeroColuna = planilha.max_column
    NumeroLinha = planilha.max_row
    ColunasDosPostos = [x for x in range(8, NumeroColuna - 1)]
    for colunas in ColunasDosPostos:
        PostosVazoes.append(planilha.cell(row = 1, column = colunas).value)
        for linhas in range(6,35):
             PostosVazoes.append(planilha.cell(row = linhas, column = colunas).value)
        PostosVazoes.append('\n')

ACOMPHPRONTO = open("C:\\Users\\Joseeustaquio\\Downloads\\PMO_deck_preliminar\\DEC_ONS_072018_RV0_VE\\DADGER1.RV0", "w") # um novo arquivo do DADGER.RVO passa a ser criado
ACOMPHPRONTO.writelines(PostosVazoes)
ACOMPHPRONTO.close()