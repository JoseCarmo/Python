import openpyxl 
from openpyxl import Workbook

planilhas = openpyxl.load_workbook('/home/jose/Downloads/ACOMPH.xlsx')
nomes_das_abas = planilhas.get_sheet_names()

colunas2 = 1

Resultados = planilhas.create_sheet("ResumoVazao") 

for abas in nomes_das_abas:
    planilha = planilhas.get_sheet_by_name(abas)
    NumeroColuna = planilha.max_column
    NumeroLinha = planilha.max_row
    ColunasDosPostos = []
    acumulador = 9
    while acumulador <= NumeroColuna:
        ColunasDosPostos.append(acumulador)
        acumulador += 8
    for colunas in ColunasDosPostos:
        Resultados.cell(row = 1, column = colunas2).value = planilha.cell(row = 1, column = colunas).value
        if planilha.cell(row = 1, column = colunas).value == (63 or 34 or 245 or 246 or 266):
            for linhas in range(6,35):
                Resultados.cell(row = linhas - 3, column = colunas2).value = planilha.cell(row = linhas, column = colunas - 1).value
        else:
            for linhas in range(6,35):
                Resultados.cell(row = linhas - 3, column = colunas2).value = planilha.cell(row = linhas, column = colunas - 1).value 
        colunas2 += 1

planilhas.save('ACOMPH_NOVO.xlsx')
planilhas.close()
